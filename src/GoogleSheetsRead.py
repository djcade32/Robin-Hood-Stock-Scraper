from typing import Dict
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from collections import Counter
import robin_stocks.robinhood as rb
from openpyxl.styles.borders import BORDER_THIN
import CreatePieCharts


# Format rule for currency
FORMAT_CURRENCY_USD_SIMPLE = '"$"#,##0.00_-'
NUMBER_WITH_DECIMAL = '#,##0.00'

# Format for thin border around cells
BORDER_FORMAT = Border(top = Side(border_style='thin', color='FF000000'),    
                              right = Side(border_style='thin', color='FF000000'), 
                              bottom = Side(border_style='thin', color='FF000000'),
                              left = Side(border_style='thin', color='FF000000'))
# Format to change title cell color to blue
TITLE_COLOR_BLUE = PatternFill(start_color='2a84fa',
                   end_color='2a84fa',
                   fill_type='solid')

COLUMN_WIDTHS = [10, 10, 10, 10, 10, 10, 10, 10, 10]

# This populates the given google sheet with the given stock info
def populate_table (workbookPath, collected_stock_info) :
    # Loading workbook and making first sheet active
    try :
        wb = load_workbook(workbookPath)
        ws = wb.active
    except Exception as ex:
        print("Error: Cannot open workbook")
        template = "An exception of type {0} occurred. Arguments:\n{1!r}"
        message = template.format(type(ex).__name__, ex.args)
        print(message)

    # Clearing the sheet
    startRowNum = 1
    amountOfStocks = len(collected_stock_info)
    listOfSectors = []
    try :
        # ws.delete_rows(startRowNum, 1000)
        wb.remove(worksheet=wb[wb.active.title])
        wb.create_sheet("Overview")
        ws = wb.active
        print("Old sheet deleted and new sheet")

        titles = ["Symbol", 
            "Sector", 
            "Price", 
            "Avg Cost", 
            "Number of Shares", 
            "Amount Spent", 
            "Closing Price", 
            "50-Day Moving Avg", 
            "200-Day Moving Avg"
            ]
        _create_titles(sheet=ws, titles=titles, color=TITLE_COLOR_BLUE)
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 20
        print("Titles appended")

        amount_spent_total = 0

        for stock in collected_stock_info :
            insertRow = [stock.symbol, 
                        stock.sector, 
                        stock.price, 
                        stock.average_buy_price, 
                        stock.quantity, 
                        stock.amount_spent, 
                        stock.closing_price, 
                        stock.fifty_day_avg, 
                        stock.twohundred_day_avg
                        ]
            listOfSectors.append(stock.sector)
            amount_spent_total = amount_spent_total + stock.amount_spent
            ws.append(insertRow)
            _pick_cell_color(startRowNum + 1, len(insertRow), stock, ws)
            startRowNum = startRowNum + 1

        portfolio_equity = float(rb.account.load_phoenix_account(info="portfolio_equity")["amount"])
        _create_statistic(sheet=ws, cell="E" + str(amountOfStocks + 3), cellStatistic="F" + str(amountOfStocks + 3), title="Total Amount Spent", statistic=amount_spent_total)
        _create_statistic(sheet=ws, cell="E" + str(amountOfStocks + 4), cellStatistic="F" + str(amountOfStocks + 4), title="Total Portfolio Equity", statistic=portfolio_equity)
        _create_statistic(sheet=ws, cell="E" + str(amountOfStocks + 5), cellStatistic="F" + str(amountOfStocks + 5), title="Profit", statistic=portfolio_equity - amount_spent_total)
        ws["F" + str(amountOfStocks + 5)].style = "Good"
        _populate_sector_chart(sheet=ws, sectors=listOfSectors, amountOfStocks=amountOfStocks)
        print("Rows updated")

        # Format range of cells to be currency
        range1 = ws["C2" : "D" + str(amountOfStocks + 1)]
        range2 = ws["F2" : "I" + str(amountOfStocks + 5)]
        range3 = ws["E2" : "E" + str(amountOfStocks + 1)]
        _format_cells(range3, format=NUMBER_WITH_DECIMAL)
        _format_cells(range1, format=FORMAT_CURRENCY_USD_SIMPLE)
        _format_cells(range2, format=FORMAT_CURRENCY_USD_SIMPLE)

        CreatePieCharts.create_pie_chart(sheet=ws, numOfStocks=amountOfStocks, sectors=listOfSectors)
        print("Charts created")
    except Exception as ex :
        print("Error: Cannot populate table")
        ws.delete_rows(startRowNum, 1000)
        template = "An exception of type {0} occurred. Arguments:\n{1!r}"
        message = template.format(type(ex).__name__, ex.args)
        print(message)
        if(ex == "PermissionError") :
            print("Make sure excel sheet is closed before running program")
    finally :
        wb.save(workbookPath)


# ------------ Helper Functions ---------------

def _create_titles(sheet, titles, color) :
    sheet.append(titles)
    for col in range(1, len(titles) + 1) :
        sheet.cell(row=1, column=col).fill = color
        sheet.cell(row=1, column=col).font = Font(bold=True)

# Creates the total statistics at the bottom of the charts
def _create_statistic(sheet, cell, cellStatistic, title, statistic) :
    sheet[cell] = title
    sheet[cell].font = Font(bold=True)
    sheet[cellStatistic] = statistic

# Creates the sector chart
def _populate_sector_chart(sheet, sectors, amountOfStocks) :
    dictSector = Counter(sectors)
    rowNum = amountOfStocks + 7
    lastStockCell = amountOfStocks + 1
    range = sheet["G" + str(rowNum) : "I" + str(rowNum + len(dictSector))]
    _border_cells(range= range)
    
    sheet.cell(row= rowNum, column= 7, value= "Sector").font = Font(bold=True)
    sheet.cell(row= rowNum, column= 8, value= "Stocks In Each Sector").font = Font(bold=True)
    sheet.cell(row= rowNum, column= 9, value= "Shares In Each Sector").font = Font(bold=True)

    i = rowNum + 1
    for sector in dictSector.items() :
        stocksInEachSector = '=COUNTIF(B2:B{lastStockCell}, "{sector}")'.format(lastStockCell= lastStockCell, sector= sector[0])
        sharesInEachSector = '=SUMIF(B2:B{lastStockCell},G{row},E2:E{lastStockCell})'.format(lastStockCell= lastStockCell, row= i)
        sheet.cell(row= i, column= 7, value= sector[0])
        sheet.cell(row= i, column= 8, value= stocksInEachSector)
        sheet.cell(row= i, column= 9, value= sharesInEachSector).number_format = NUMBER_WITH_DECIMAL
        i = i + 1
        
    sheet["G" + str(i + 1)] = "Total of Stocks"
    sheet["G" + str(i + 1)].font = Font(bold=True)
    sheet.cell(row= i + 1, column= 8, value= '=SUM(H{startCell}:H{endCell})'.format(startCell= rowNum + 1, endCell= rowNum + 1 + len(dictSector)))

    sheet["G" + str(i + 2)] = "Total of Shares"
    sheet["G" + str(i + 2)].font = Font(bold=True)
    sheet.cell(row= i + 2, column= 8, value= '=SUM(I27:I34)').number_format = NUMBER_WITH_DECIMAL

# Put thin border around specified range of cells
def _border_cells(range) :
    for cellTuple in range :
        for cell in cellTuple :
            cell.border = BORDER_FORMAT

# Formats the given range of cells with the given format rule
def _format_cells(range, format) :
    for cellTuple in range :
        for cell in cellTuple :
            cell.number_format = format

# Determines the color of the cell based off of the stocks 50
# and 200 day moving avg
def _pick_cell_color(rowNum, columnNum, stock, sheet) :
    if stock.fifty_day_avg > stock.price :
        sheet.cell(row=rowNum, column=columnNum-1).style = "Bad"
    else :
        sheet.cell(row=rowNum, column=columnNum-1).style = "Good"
    if stock.twohundred_day_avg > stock.price :
        sheet.cell(row=rowNum, column=columnNum).style = "Bad"
    else :
        sheet.cell(row=rowNum, column=columnNum).style = "Good"
